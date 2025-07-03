import streamlit as st
import openpyxl
from io import BytesIO
import traceback
import os
from copy import copy
from openpyxl.cell.cell import MergedCell # Importante para la detección

st.set_page_config(page_title="Consolidador Robusto", layout="wide")
st.title("Consolidador REM (Método Seguro)")

st.warning(
    "**Importante:** Este método reconstruye el archivo final desde cero para máxima compatibilidad. "
    "Se copiarán valores, formatos, anchos/altos de celda y celdas combinadas. **Gráficos, imágenes y tablas dinámicas no serán transferidos.**"
)

# --- Funciones de Ayuda para Copiar ---
def copy_sheet_properties(source_ws, target_ws):
    """Copia propiedades de la hoja como anchos de columna, altos de fila y celdas combinadas."""
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

def copy_cell(source_cell, target_cell):
    """Copia valor y todos los estilos de una celda a otra."""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

# --- Interfaz de Streamlit ---
uploaded_files = st.file_uploader(
    "Selecciona los archivos Excel (el primero será la plantilla)",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True
)

if 'processed_file' not in st.session_state:
    st.session_state.processed_file = None
    st.session_state.file_name = None

if uploaded_files:
    if st.button("✨ Consolidar Archivos (Método Seguro)"):
        with st.spinner("Iniciando proceso..."):
            try:
                # --- PASO 1: Sumar todos los datos numéricos ---
                sumas_consolidadas = {}
                progress_bar = st.progress(0, text="Paso 1/3: Sumando datos de los archivos...")
                # ... (El código de suma no cambia y es correcto)
                for i, data_file in enumerate(uploaded_files):
                    data_file.seek(0)
                    wb_data = openpyxl.load_workbook(data_file, data_only=True)
                    for hoja_nombre in wb_data.sheetnames:
                        if hoja_nombre not in sumas_consolidadas:
                            sumas_consolidadas[hoja_nombre] = {}
                        ws_data = wb_data[hoja_nombre]
                        for fila in ws_data.iter_rows():
                            for celda in fila:
                                if isinstance(celda.value, (int, float)):
                                    ref = celda.coordinate
                                    sumas_consolidadas[hoja_nombre][ref] = sumas_consolidadas[hoja_nombre].get(ref, 0) + celda.value
                    progress_bar.progress((i + 1) / len(uploaded_files), text=f"Paso 1/3: Leyendo {data_file.name}")

                # --- PASO 2: Construir un nuevo libro copiando la plantilla ---
                progress_bar.progress(0, text="Paso 2/3: Reconstruyendo la plantilla desde cero...")
                template_file = uploaded_files[0]
                template_file.seek(0)
                wb_template = openpyxl.load_workbook(template_file)
                wb_final = openpyxl.Workbook()
                wb_final.remove(wb_final.active)

                for i, hoja_nombre in enumerate(wb_template.sheetnames):
                    source_ws = wb_template[hoja_nombre]
                    target_ws = wb_final.create_sheet(title=hoja_nombre)
                    
                    # Primero, copiar la estructura (incluidas las celdas combinadas)
                    copy_sheet_properties(source_ws, target_ws)

                    # Segundo, copiar el contenido y estilo de cada celda individual
                    for fila in source_ws.iter_rows():
                        for celda in fila:
                            # --- LA CORRECCIÓN CLAVE ESTÁ AQUÍ ---
                            # Si la celda es un marcador de posición de una celda combinada, la omitimos.
                            # La estructura ya fue creada por 'merge_cells'
                            if isinstance(celda, MergedCell):
                                continue
                            
                            new_cell = target_ws.cell(row=celda.row, column=celda.column)
                            copy_cell(celda, new_cell)
                    
                    progress_bar.progress((i + 1) / len(wb_template.sheetnames), text=f"Paso 2/3: Copiando hoja '{hoja_nombre}'...")

                # --- PASO 3: Escribir los datos sumados ---
                progress_bar.progress(0, text="Paso 3/3: Escribiendo valores consolidados...")
                for hoja_nombre, celdas in sumas_consolidadas.items():
                    if hoja_nombre in wb_final.sheetnames:
                        ws_final = wb_final[hoja_nombre]
                        for celda_ref, valor_sumado in celdas.items():
                            try:
                                ws_final[celda_ref].value = valor_sumado
                            except Exception:
                                pass
                
                # --- PASO 4: Guardar y preparar para la descarga ---
                progress_bar.progress(1.0, text="Generando archivo final...")
                output = BytesIO()
                wb_final.save(output)
                output.seek(0)
                
                base_name = os.path.splitext(template_file.name)[0]
                final_filename = f"consolidado_seguro_{base_name}.xlsx"
                
                st.session_state.processed_file = output
                st.session_state.file_name = final_filename
                progress_bar.empty()

            except Exception as e:
                st.error(f"Ocurrió un error crítico durante el proceso: {e}")
                st.error(traceback.format_exc())
                st.session_state.processed_file = None

# --- Lógica de descarga (sin cambios) ---
if st.session_state.processed_file is not None:
    st.success("✅ ¡Consolidación completada! El archivo reconstruido está listo.")
    st.download_button(
        label="📥 Descargar archivo consolidado (.xlsx)",
        data=st.session_state.processed_file,
        file_name=st.session_state.file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    if st.button("Limpiar y empezar de nuevo"):
        st.session_state.processed_file = None
        st.session_state.file_name = None
        st.rerun()
